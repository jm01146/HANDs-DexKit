"""
Magnetometer Data Logger with 3D Printer Control
-------------------------------------------------
Hardware:
  - Teensy  → SENSOR_PORT  (outputs "x,y,z" float lines from QMC5883)
  - Printer → PRINTER_PORT (G-code over USB, moves Z axis only)

Key bindings (no Enter needed):
  M  →  move printer to next Z position in POSITION_LIST_MM
         (wraps back to the first position after the last)
  R  →  record 1000 samples at current Z, save to Excel
  H  →  home printer (G28)
  Q  →  save and quit
  ?  →  show help

Rules:
  - Movement is locked while a recording is in progress.
  - Sessions accumulate across runs (workbook is appended, not overwritten).
"""

import os
import sys
import time
import threading
from datetime import datetime

import serial
from serial import SerialException
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

# ── USER CONFIG ───────────────────────────────────────────────────────────────
SENSOR_PORT         = "COM4"        # Teensy
PRINTER_PORT        = "COM3"        # 3D printer
BAUD_SENSOR         = 115200
BAUD_PRINTER        = 115200

OUTPUT_XLSX         = "Zaxis5_25Distance.xlsx"
SAMPLES_PER_SESSION = 1000
FEEDRATE            = 1000          # mm/min for Z moves
HOME_ON_START       = False

# Axis the printer moves on — prompted on launch, or hardcode "X"/"Y"/"Z" to skip prompt.
MOVE_AXIS: str | None = None

# Ordered list of positions (mm) to cycle through with M key.
POSITION_LIST_MM = [0, 5, 7, 9, 11, 13, 15, 17, 19, 21, 23, 25]

# Key bindings (single character, case-insensitive)
KEY_MOVE   = "m"
KEY_RECORD = "r"
KEY_HOME   = "h"
KEY_QUIT   = "q"

# ── CROSS-PLATFORM SINGLE-KEYPRESS ────────────────────────────────────────────
if sys.platform == "win32":
    import msvcrt

    def _getch():
        return msvcrt.getwch()

else:
    import tty, termios

    def _getch():
        fd  = sys.stdin.fileno()
        old = termios.tcgetattr(fd)
        try:
            tty.setraw(fd)
            return sys.stdin.read(1)
        finally:
            termios.tcsetattr(fd, termios.TCSADRAIN, old)


# ── PRINTER ───────────────────────────────────────────────────────────────────
class Printer:
    # Making sure that the serial port is cleaned at start up of the code and connected properly
    def __init__(self, port: str, baud: int):
        self._ser  = serial.Serial(port, baud, timeout=2)
        self._lock = threading.Lock()
        time.sleep(2)
        self._ser.reset_input_buffer()
        self._flush_startup()
# The actual flushing function of the code that will be used in __init__ (that is the start up method)
    def _flush_startup(self):
        deadline = time.time() + 3
        while time.time() < deadline:
            line = self._ser.readline().decode("utf-8", errors="replace").strip()
            if line:
                print(f"  [printer] {line}")
# How we will send instructions to the printer to move
    def _send(self, cmd: str):
        self._ser.write((cmd.strip() + "\n").encode())
# We lock the python code until the printer says its okay to send another command 
    def _wait_ok(self, timeout: float = 60.0):
        deadline = time.time() + timeout
        while time.time() < deadline:
            raw = self._ser.readline()
            if not raw:
                continue
            reply = raw.decode("utf-8", errors="replace").strip()
            if reply:
                print(f"  [printer] {reply}")
            if reply.lower().startswith("ok"):
                return True
        raise TimeoutError("Printer did not reply 'ok' in time.")
# Home command for the printer to avoid writing it constantly
    def home(self):
        with self._lock:
            print("  Homing...")
            self._send("G1 Z0 F1000")
            self._wait_ok(timeout=120)
            print("  Home complete.")
# Set the axis you want to move in main and then how far you want to move per session 
    def move_axis(self, axis: str, position_mm: float):
        """Move one axis to an absolute position and block until physically complete."""
        with self._lock:
            self._send("G90")
            self._wait_ok()
            self._send(f"G1 {axis.upper()}{position_mm:.3f} F{FEEDRATE}")
            self._wait_ok(timeout=60)
            self._send("M400")
            self._wait_ok(timeout=60)
        print(f"  Printer at {axis.upper()}={position_mm:.3f} mm")

    def close(self):
        self._ser.close()


# ── SENSOR (Teensy) ───────────────────────────────────────────────────────────
class SensorReader:
    def __init__(self, port: str, baud: int):
        self._ser     = serial.Serial(port, baud, timeout=1)
        self._lock    = threading.Lock()
        self._latest  = None          # (x, y, z) floats
        self._running = True
        time.sleep(1.5)
        self._ser.reset_input_buffer()
        threading.Thread(target=self._loop, daemon=True).start()

    def _parse(self, line: str) -> tuple | None:
        """Extract exactly 3 floats from a line like '1.23,-4.56,78.9'."""
        cleaned = "".join(c if (c.isdigit() or c in "-., ") else " " for c in line)
        parts   = [p for p in cleaned.replace(",", " ").split() if p]
        nums    = []
        for p in parts:
            try:
                nums.append(float(p))
            except ValueError:
                continue
            if len(nums) == 3:
                break
        return tuple(nums) if len(nums) == 3 else None

    def _loop(self):
        while self._running:
            try:
                raw = self._ser.readline()
            except SerialException:
                break
            if not raw:
                continue
            parsed = self._parse(raw.decode("utf-8", errors="replace").strip())
            if parsed:
                with self._lock:
                    self._latest = parsed

    def latest(self):
        with self._lock:
            return self._latest

    def collect(self, n: int) -> list:
        """Block until n distinct samples collected. Returns list of (x,y,z)."""
        samples = []
        prev    = None
        print(f"  Recording {n} samples ", end="", flush=True)
        while len(samples) < n:
            s = self.latest()
            if s is not None and s is not prev:
                samples.append(s)
                prev = s
                if len(samples) % 100 == 0:
                    print(".", end="", flush=True)
            else:
                time.sleep(0.001)
        print(" done")
        return samples

    def stop(self):
        self._running = False
        self._ser.close()


# ── EXCEL ─────────────────────────────────────────────────────────────────────
RAW_HEADERS = [
    "session_id", "timestamp_iso", "epoch_s",
    "axis", "distance_mm", "sample_index",
    "Bx_uT", "By_uT", "Bz_uT",
]


_HDR_FILL = PatternFill("solid", start_color="2F4F8F")
_HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_CENTER   = Alignment(horizontal="center")


def _style_headers(ws):
    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment = _HDR_FILL, _HDR_FONT, _CENTER


def open_or_create_workbook(path: str):
    if os.path.exists(path):
        wb      = load_workbook(path)
        ws_raw  = wb["raw_samples"]
        ws_stat = wb["session_stats"]
    else:
        wb = Workbook()

        ws_raw       = wb.active
        ws_raw.title = "raw_samples"
        ws_raw.append(RAW_HEADERS)
        _style_headers(ws_raw)
        for i, w in enumerate([10, 24, 14, 6, 14, 12, 12, 12, 12], 1):
            ws_raw.column_dimensions[get_column_letter(i)].width = w


    return wb, ws_raw



def append_session(ws_raw, session_id: int,
                   axis: str, position_mm: float, samples: list):
    ts  = datetime.now().isoformat(timespec="milliseconds")
    now = time.time()

    for i, (bx, by, bz) in enumerate(samples):
        ws_raw.append([session_id, ts, now, axis, position_mm, i, bx, by, bz])



# ── STATE ─────────────────────────────────────────────────────────────────────
class State:
    def __init__(self, axis: str):
        self.axis       = axis.upper()
        self.pos_index  = 0
        self.position   = POSITION_LIST_MM[0]
        self.recording  = threading.Event()

    @property
    def next_index(self):
        return (self.pos_index + 1) % len(POSITION_LIST_MM)

    def advance(self):
        self.pos_index = self.next_index
        self.position  = POSITION_LIST_MM[self.pos_index]
        return self.pos_index


# ── ACTIONS ───────────────────────────────────────────────────────────────────
def action_move(printer: Printer, state: State):
    if state.recording.is_set():
        print("\n  Recording in progress — movement locked.\n")
        return

    new_idx  = state.advance()
    pos      = state.position
    total    = len(POSITION_LIST_MM)
    wrapping = (new_idx == 0)
    tag      = "  (back to start)" if wrapping else ""
    print(f"\n  [{new_idx + 1}/{total}] Moving {state.axis} to {pos} mm{tag} ...")

    def _move():
        try:
            printer.move_axis(state.axis, pos)
            nxt = POSITION_LIST_MM[state.next_index]
            print(f"  Next M → {state.axis}={nxt} mm  |  Press R to record here.\n")
        except (SerialException, TimeoutError) as e:
            print(f"  Printer error: {e}\n")

    threading.Thread(target=_move, daemon=True).start()


def action_record(sensor: SensorReader, state: State,
                  wb, ws_raw,
                  session_counter: list, xlsx_path: str):
    if state.recording.is_set():
        print("\n  Already recording — wait for it to finish.\n")
        return

    state.recording.set()

    def _run():
        try:
            sid = session_counter[0]
            pos = state.position
            print(f"\n  Session {sid} | {state.axis}={pos} mm")
            samples = sensor.collect(SAMPLES_PER_SESSION)
            append_session(ws_raw, sid, state.axis, pos, samples)
            wb.save(xlsx_path)
            print(f"  Session {sid} saved → {xlsx_path}\n")
            session_counter[0] += 1
        finally:
            state.recording.clear()

    threading.Thread(target=_run, daemon=True).start()


# ── MAIN ──────────────────────────────────────────────────────────────────────
def print_help(state: State):
    total = len(POSITION_LIST_MM)
    cur   = state.position
    nxt   = POSITION_LIST_MM[state.next_index]
    print(f"""
  ┌─────────────────────────────────────────────────────┐
  │  Key    Action                                      │
  │  ─────────────────────────────────────────────────  │
  │  {KEY_MOVE.upper()}      Move {state.axis} to next position (cycles)    │
  │  {KEY_RECORD.upper()}      Record {SAMPLES_PER_SESSION} samples at current pos        │
  │  {KEY_HOME.upper()}      Home printer (G28)                   │
  │  {KEY_QUIT.upper()}      Save and quit                         │
  │  ?      Show this help                              │
  └─────────────────────────────────────────────────────┘
  Axis: {state.axis}   |   Positions ({total}): {POSITION_LIST_MM}
  Current: {cur} mm   |   Next M → {state.axis}={nxt} mm
""")


def main():
    print("=== Magnetometer + Printer Logger ===")
    print(f"Sensor  : {SENSOR_PORT}  @ {BAUD_SENSOR}  (Teensy / QMC5883)")
    print(f"Printer : {PRINTER_PORT} @ {BAUD_PRINTER}")
    print(f"Output  : {OUTPUT_XLSX}\n")

    # Resolve axis — use config value or prompt
    axis = MOVE_AXIS
    if axis is None:
        while True:
            raw = input("  Which axis will the printer move on? [X/Y/Z]: ").strip().upper()
            if raw in ("X", "Y", "Z"):
                axis = raw
                break
            print("  Please enter X, Y, or Z.")
    else:
        axis = axis.upper()
        if axis not in ("X", "Y", "Z"):
            print(f"  Invalid MOVE_AXIS '{axis}' in config. Must be X, Y, or Z.")
            return
        print(f"  Axis: {axis} (from config)")
    print()

    wb, ws_raw = open_or_create_workbook(OUTPUT_XLSX)

    print("Connecting to printer...")
    try:
        printer = Printer(PRINTER_PORT, BAUD_PRINTER)
    except SerialException as e:
        print(f"Could not open printer port: {e}")
        return

    print("Connecting to Teensy sensor...")
    try:
        sensor = SensorReader(SENSOR_PORT, BAUD_SENSOR)
    except SerialException as e:
        print(f"Could not open sensor port: {e}")
        printer.close()
        return

    if HOME_ON_START:
        printer.home()

    state           = State(axis)
    session_counter = [1]

    print_help(state)
    print("Press a key...\n")

    try:
        while True:
            key = _getch().lower()

            if key == KEY_QUIT:
                print("\nQuitting...")
                break

            elif key == KEY_MOVE:
                action_move(printer, state)

            elif key == KEY_RECORD:
                action_record(sensor, state,
                              wb, ws_raw,
                              session_counter, OUTPUT_XLSX)

            elif key == KEY_HOME:
                if state.recording.is_set():
                    print("\n  Recording in progress — movement locked.\n")
                else:
                    print("\n  Homing...")
                    threading.Thread(target=printer.home, daemon=True).start()

            elif key == "?":
                print_help(state)

    except KeyboardInterrupt:
        pass

    finally:
        wb.save(OUTPUT_XLSX)
        sensor.stop()
        printer.close()
        total = session_counter[0] - 1
        print(f"\nDone. {total} session(s) recorded → {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
